"""Generate synthetic ESG test data files"""
import random
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows


def generate_cement_plant_a():
    """Generate cement_plant_a.xlsx with multiple sheets"""
    wb = Workbook()
    wb.remove(wb.active)
    
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    
    # Energy Sheet
    ws_energy = wb.create_sheet("Energy")
    ws_energy.merge_cells('A1:B1')
    ws_energy['A1'] = "Cement Plant A - Energy Data 2024"
    ws_energy['A1'].font = Font(bold=True, size=14)
    
    ws_energy['A3'] = "Month"
    ws_energy['B3'] = "Electricity Consumption (kWh)"
    ws_energy['C3'] = "Power Usage (MWh)"
    ws_energy['D3'] = "Energy - Electrical (GJ)"
    
    for i, month in enumerate(months, start=4):
        production = random.randint(8000, 12000)
        ws_energy[f'A{i}'] = month
        ws_energy[f'B{i}'] = random.randint(500, 1500) * production
        ws_energy[f'C{i}'] = f"=B{i}/1000"
        ws_energy[f'D{i}'] = f"=C{i}*3.6"
    
    ws_energy['A16'] = "Total"
    ws_energy['A16'].font = Font(bold=True)
    ws_energy['B16'] = "=SUM(B4:B15)"
    ws_energy['C16'] = "=SUM(C4:C15)"
    ws_energy['D16'] = "=SUM(D4:D15)"
    
    # Emissions Sheet
    ws_emissions = wb.create_sheet("Emissions")
    ws_emissions.merge_cells('A1:C1')
    ws_emissions['A1'] = "Cement Plant A - Emissions Data 2024"
    ws_emissions['A1'].font = Font(bold=True, size=14)
    
    ws_emissions['A3'] = "Month"
    ws_emissions['B3'] = "CO2 Emissions (tonnes)"
    ws_emissions['C3'] = "Carbon Dioxide (kg)"
    ws_emissions['D3'] = "GHG - Scope 1 (tCO2e)"
    ws_emissions['E3'] = "Production (tonnes clinker)"
    ws_emissions['F3'] = "Intensity (kg CO2/tonne)"
    
    for i, month in enumerate(months, start=4):
        production = random.randint(8000, 12000)
        intensity = random.randint(800, 1100)
        emissions = production * intensity / 1000
        
        ws_emissions[f'A{i}'] = month
        ws_emissions[f'B{i}'] = round(emissions, 2)
        ws_emissions[f'C{i}'] = f"=B{i}*1000"
        ws_emissions[f'D{i}'] = f"=B{i}"
        ws_emissions[f'E{i}'] = production
        ws_emissions[f'F{i}'] = f"=B{i}*1000/E{i}"
    
    ws_emissions['A16'] = "Total"
    ws_emissions['A16'].font = Font(bold=True)
    ws_emissions['B16'] = "=SUM(B4:B15)"
    ws_emissions['E16'] = "=SUM(E4:E15)"
    ws_emissions['F16'] = "=B16*1000/E16"
    
    # Water Sheet
    ws_water = wb.create_sheet("Water")
    ws_water['A1'] = "Month"
    ws_water['B1'] = "Water Consumption (m³)"
    ws_water['C1'] = "Water Withdrawal (liters)"
    ws_water['D1'] = "Recycled Water (%)"
    
    for i, month in enumerate(months, start=2):
        ws_water[f'A{i}'] = month
        ws_water[f'B{i}'] = random.randint(15000, 25000)
        ws_water[f'C{i}'] = f"=B{i}*1000"
        ws_water[f'D{i}'] = random.randint(20, 45)
    
    ws_water['A14'] = "Total"
    ws_water['B14'] = "=SUM(B2:B13)"
    ws_water['D14'] = "=AVERAGE(D2:D13)"
    
    output_path = Path("data/sample-inputs/cement_plant_a.xlsx")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Created: {output_path}")


def generate_steel_facility_b():
    """Generate steel_facility_b.csv with 24 months of data"""
    months = []
    for year in [2023, 2024]:
        for month in range(1, 13):
            months.append(f"{year}-{month:02d}")
    
    data = []
    for month in months:
        production = random.randint(50000, 80000)
        fuel_gj = production * random.uniform(18, 25)
        emissions_kg = production * random.randint(1800, 2500)
        
        data.append({
            "Month": month,
            "Production (tonnes)": production,
            "Fuel Consumption (GJ)": round(fuel_gj, 2),
            "Emissions (kg CO2)": emissions_kg,
            "Energy Intensity (GJ/tonne)": round(fuel_gj / production, 3),
            "Carbon Intensity (kg CO2/tonne)": round(emissions_kg / production, 2)
        })
    
    df = pd.DataFrame(data)
    output_path = Path("data/sample-inputs/steel_facility_b.csv")
    df.to_csv(output_path, index=False)
    print(f"Created: {output_path}")


def generate_messy_data():
    """Generate messy_data.xlsx with edge cases"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Main Data"
    
    ws.merge_cells('A1:E1')
    ws['A1'] = "COMPANY LOGO"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:E2')
    ws['A2'] = "Environmental Performance Report"
    ws['A2'].font = Font(size=12)
    ws['A2'].alignment = Alignment(horizontal='center')
    
    ws['A4'] = "Facility: Manufacturing Plant XYZ"
    ws['A5'] = "Reporting Period: 2024"
    
    ws['A8'] = "Month"
    ws['B8'] = "Energy (kWh)"
    ws['C8'] = "Emissions (tonnes CO2)"
    ws['D8'] = "Water (m³)"
    ws['E8'] = "Waste (kg)"
    ws['F8'] = "Intensity"
    
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    
    for i, month in enumerate(months, start=9):
        ws[f'A{i}'] = month
        ws[f'B{i}'] = random.randint(100000, 500000) if random.random() > 0.1 else None
        ws[f'C{i}'] = random.randint(50, 200)
        ws[f'D{i}'] = random.randint(5000, 15000) if random.random() > 0.15 else None
        ws[f'E{i}'] = random.randint(1000, 5000)
        
        if i == 11:
            ws[f'F{i}'] = "=C{i}/0"
        else:
            ws[f'F{i}'] = f"=C{i}/B{i}*1000000" if ws[f'B{i}'].value else None
    
    ws_validation = wb.create_sheet("_Validation")
    ws_validation['A1'] = "Valid Metrics"
    ws_validation['A2'] = "Energy"
    ws_validation['A3'] = "Emissions"
    ws_validation['A4'] = "Water"
    ws_validation['A5'] = "Waste"
    ws_validation.sheet_state = 'hidden'
    
    output_path = Path("data/sample-inputs/messy_data.xlsx")
    wb.save(output_path)
    print(f"Created: {output_path}")


def generate_large_dataset():
    """Generate large_facility_data.csv with 5 years of daily data"""
    print("Generating large dataset (5 years of daily data)...")
    
    data = []
    start_date = pd.Timestamp("2020-01-01")
    
    for day in range(365 * 5):
        date = start_date + pd.Timedelta(days=day)
        production = random.randint(800, 1200)
        
        data.append({
            "Date": date.strftime("%Y-%m-%d"),
            "Production (tonnes)": production,
            "Electricity (kWh)": production * random.randint(600, 900),
            "Natural Gas (m³)": production * random.randint(100, 200),
            "Water (m³)": production * random.uniform(1.5, 3.0),
            "CO2 Emissions (kg)": production * random.randint(850, 1050),
            "NOx Emissions (kg)": production * random.uniform(0.5, 1.5),
            "SOx Emissions (kg)": production * random.uniform(0.3, 0.8),
            "Waste Generated (kg)": production * random.randint(50, 150),
            "Waste Recycled (kg)": production * random.randint(30, 100),
            "Employees on Site": random.randint(150, 200),
            "Safety Incidents": random.choices([0, 0, 0, 0, 1], weights=[85, 10, 3, 1, 1])[0]
        })
    
    df = pd.DataFrame(data)
    output_path = Path("data/sample-inputs/large_facility_data.csv")
    df.to_csv(output_path, index=False)
    print(f"Created: {output_path} ({len(df)} rows)")


def generate_mixed_units():
    """Generate mixed_units.xlsx with intentionally inconsistent units"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Energy Data"
    
    ws['A1'] = "Facility"
    ws['B1'] = "Energy Consumption"
    ws['C1'] = "Unit"
    ws['D1'] = "Emissions"
    ws['E1'] = "Unit"
    
    facilities = [
        ("Plant A", 1500000, "kWh", 750, "tonnes CO2"),
        ("Plant B", 1.5, "GWh", 750000, "kg CO2"),
        ("Plant C", 5400, "GJ", 0.75, "kt CO2"),
        ("Plant D", 5400000, "MJ", 750, "t CO2"),
        ("Plant E", 1500, "MWh", 7.5e5, "g CO2"),
    ]
    
    for i, (facility, energy, e_unit, emissions, em_unit) in enumerate(facilities, start=2):
        ws[f'A{i}'] = facility
        ws[f'B{i}'] = energy
        ws[f'C{i}'] = e_unit
        ws[f'D{i}'] = emissions
        ws[f'E{i}'] = em_unit
    
    output_path = Path("data/sample-inputs/mixed_units.xlsx")
    wb.save(output_path)
    print(f"Created: {output_path}")


if __name__ == "__main__":
    print("Generating synthetic ESG test data...")
    generate_cement_plant_a()
    generate_steel_facility_b()
    generate_messy_data()
    generate_large_dataset()
    generate_mixed_units()
    print("\nAll test data files generated successfully!")
