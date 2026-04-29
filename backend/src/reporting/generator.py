"""Report generation service."""

import logging
from pathlib import Path
from typing import Dict, List
from uuid import UUID

logger = logging.getLogger(__name__)


class ReportGenerator:
    """Generate ESG reports with PDF conversion."""

    def __init__(self, templates_dir: str, db_session=None):
        self.templates_dir = Path(templates_dir)
        self.db = db_session

    def convert_to_pdf(self, docx_path: str) -> str:
        """Convert DOCX to PDF.
        
        Args:
            docx_path: Path to DOCX file
            
        Returns:
            Path to PDF file
        """
        from docx2pdf import convert
        
        pdf_path = docx_path.replace('.docx', '.pdf')
        convert(docx_path, pdf_path)
        
        return pdf_path

    def generate_brsr_excel(self, upload_id: UUID, company_info: Dict, normalized_data: List[Dict]) -> str:
        """Generate BRSR in Excel format.
        
        Args:
            upload_id: Upload UUID
            company_info: Company information
            normalized_data: Normalized data records
            
        Returns:
            Path to Excel file
        """
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        
        wb = openpyxl.Workbook()
        wb.create_sheet('Section A')
        wb.create_sheet('Principle 6')
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Section A
        ws_a = wb['Section A']
        ws_a['A1'] = 'BRSR - Section A'
        ws_a['A1'].font = Font(bold=True, size=14)
        ws_a['A2'] = 'Company Name:'
        ws_a['B2'] = company_info.get('name', '')
        ws_a['A3'] = 'CIN:'
        ws_a['B3'] = company_info.get('cin', '')
        
        # Principle 6
        ws_p6 = wb['Principle 6']
        ws_p6['A1'] = 'Principle 6: Environment'
        ws_p6['A1'].font = Font(bold=True, size=14)
        
        headers = ['Indicator', 'Value', 'Unit', 'Original Value', 'Original Unit']
        for col, header in enumerate(headers, start=1):
            cell = ws_p6.cell(row=3, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        
        row = 4
        for record in normalized_data:
            ws_p6[f'A{row}'] = record.get('indicator', '')
            ws_p6[f'B{row}'] = record.get('normalized_value', '')
            ws_p6[f'C{row}'] = record.get('normalized_unit', '')
            ws_p6[f'D{row}'] = record.get('original_value', '')
            ws_p6[f'E{row}'] = record.get('original_unit', '')
            row += 1
        
        output_dir = Path('data/outputs')
        output_dir.mkdir(parents=True, exist_ok=True)
        output_path = output_dir / f'BRSR_Excel_{upload_id}.xlsx'
        wb.save(str(output_path))
        
        return str(output_path)

    def generate_reports(
        self,
        upload_id: UUID,
        data: Dict,
        output_dir: str,
        formats: List[str] = ["brsr"]
    ) -> List[Dict]:
        """Generate reports."""
        reports = []
        return reports
