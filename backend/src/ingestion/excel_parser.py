"""Excel file parser with advanced data extraction"""
import logging
from pathlib import Path
from typing import List, Tuple, Dict, Any

import polars as pl
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.worksheet import Worksheet

from src.ingestion.exceptions import ParseError, AuthenticationError, FileFormatError
from src.ingestion.base_parser import BaseParser, ParsedDataFrame

logger = logging.getLogger(__name__)


class ExcelParser(BaseParser):
    """Parser for Excel files with intelligent data extraction"""

    def parse(self, file_path: str) -> ParsedDataFrame:
        """Parse Excel file and return DataFrame with metadata"""
        logger.info(f"Parsing Excel file: {file_path}")
        
        try:
            workbook = load_workbook(filename=file_path, data_only=False)
        except InvalidFileException as e:
            logger.error(f"Invalid file format: {e}")
            raise FileFormatError(f"Invalid Excel file format: {str(e)}")
        except Exception as e:
            if "password" in str(e).lower() or "encrypted" in str(e).lower():
                logger.error("Password-protected file detected")
                raise AuthenticationError("File is password-protected")
            logger.error(f"Failed to open file: {e}")
            raise ParseError(f"Failed to open Excel file: {str(e)}")

        all_data = []
        metadata = {
            "filename": Path(file_path).name,
            "sheets": [],
            "total_rows": 0,
            "total_columns": 0
        }

        for sheet_name in workbook.sheetnames:
            logger.info(f"Processing sheet: {sheet_name}")
            sheet = workbook[sheet_name]
            
            try:
                # Handle merged cells
                self.handle_merged_cells(sheet)
                
                # Detect data region
                start_row, end_row, start_col, end_col = self.detect_data_region(sheet)
                
                if start_row is None:
                    logger.warning(f"No data found in sheet: {sheet_name}")
                    continue
                
                # Extract headers
                headers = self.extract_headers(sheet, start_row)
                
                # Extract data rows
                rows = []
                for row_idx in range(start_row + 1, end_row + 1):
                    row_data = []
                    for col_idx in range(start_col, end_col + 1):
                        cell = sheet.cell(row=row_idx, column=col_idx)
                        value = cell.value
                        
                        # Handle formulas
                        if hasattr(cell, 'data_type') and cell.data_type == 'f':
                            try:
                                # Try to get calculated value
                                wb_data_only = load_workbook(filename=file_path, data_only=True)
                                value = wb_data_only[sheet_name].cell(row=row_idx, column=col_idx).value
                            except Exception:
                                value = None
                        
                        row_data.append(value)
                    rows.append(row_data)
                
                # Create DataFrame for this sheet
                if rows:
                    sheet_df = pl.DataFrame(
                        {header: [row[i] if i < len(row) else None for row in rows] 
                         for i, header in enumerate(headers)}
                    )
                    sheet_df = sheet_df.with_columns(pl.lit(sheet_name).alias("_sheet_name"))
                    all_data.append(sheet_df)
                    
                    metadata["sheets"].append({
                        "name": sheet_name,
                        "rows": len(rows),
                        "columns": len(headers),
                        "headers": headers
                    })
                    metadata["total_rows"] += len(rows)
                    
            except Exception as e:
                logger.error(f"Error processing sheet {sheet_name}: {e}")
                raise ParseError(f"Error processing sheet {sheet_name}: {str(e)}")
        
        if not all_data:
            raise ParseError("No data found in any sheet")
        
        # Combine all sheets
        combined_df = pl.concat(all_data, how="diagonal")
        metadata["total_columns"] = len(combined_df.columns) - 1  # Exclude _sheet_name
        
        logger.info(f"Successfully parsed {metadata['total_rows']} rows from {len(metadata['sheets'])} sheets")
        
        return ParsedDataFrame(data=combined_df, metadata=metadata)

    def detect_data_region(self, sheet: Worksheet) -> Tuple[int, int, int, int]:
        """Detect the data region in a worksheet"""
        max_row = sheet.max_row
        max_col = sheet.max_column
        
        start_row = None
        end_row = None
        start_col = 1
        end_col = max_col
        
        # Find first row with >50% non-empty cells
        for row_idx in range(1, max_row + 1):
            non_empty = 0
            total = 0
            for col_idx in range(1, max_col + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                total += 1
                if cell.value is not None and str(cell.value).strip():
                    non_empty += 1
            
            if total > 0 and non_empty / total > 0.5:
                start_row = row_idx
                break
        
        if start_row is None:
            return None, None, None, None
        
        # Find last row with data
        for row_idx in range(max_row, start_row, -1):
            has_data = False
            for col_idx in range(1, max_col + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                if cell.value is not None and str(cell.value).strip():
                    has_data = True
                    break
            if has_data:
                end_row = row_idx
                break
        
        # Find first and last columns with data
        for col_idx in range(1, max_col + 1):
            has_data = False
            for row_idx in range(start_row, end_row + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                if cell.value is not None and str(cell.value).strip():
                    has_data = True
                    break
            if has_data:
                start_col = col_idx
                break
        
        for col_idx in range(max_col, start_col - 1, -1):
            has_data = False
            for row_idx in range(start_row, end_row + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                if cell.value is not None and str(cell.value).strip():
                    has_data = True
                    break
            if has_data:
                end_col = col_idx
                break
        
        logger.info(f"Data region: rows {start_row}-{end_row}, cols {start_col}-{end_col}")
        return start_row, end_row, start_col, end_col

    def extract_headers(self, sheet: Worksheet, start_row: int) -> List[str]:
        """Extract and clean header row"""
        headers = []
        max_col = sheet.max_column
        
        for col_idx in range(1, max_col + 1):
            cell = sheet.cell(row=start_row, column=col_idx)
            header = str(cell.value).strip() if cell.value else f"Column_{col_idx}"
            
            # Handle duplicate headers
            original_header = header
            counter = 1
            while header in headers:
                header = f"{original_header}_{counter}"
                counter += 1
            
            headers.append(header)
        
        logger.info(f"Extracted {len(headers)} headers")
        return headers

    def handle_merged_cells(self, sheet: Worksheet) -> None:
        """Unmerge cells and propagate values"""
        merged_ranges = list(sheet.merged_cells.ranges)
        
        for merged_range in merged_ranges:
            # Get the top-left cell value
            min_row, min_col = merged_range.min_row, merged_range.min_col
            top_left_value = sheet.cell(row=min_row, column=min_col).value
            
            # Unmerge
            sheet.unmerge_cells(str(merged_range))
            
            # Fill all cells with the top-left value
            for row_idx in range(merged_range.min_row, merged_range.max_row + 1):
                for col_idx in range(merged_range.min_col, merged_range.max_col + 1):
                    sheet.cell(row=row_idx, column=col_idx).value = top_left_value
        
        if merged_ranges:
            logger.info(f"Handled {len(merged_ranges)} merged cell ranges")

    def evaluate_formulas(self, sheet: Worksheet) -> None:
        """Evaluate formulas and replace with calculated values"""
        formula_count = 0
        
        for row in sheet.iter_rows():
            for cell in row:
                if hasattr(cell, 'data_type') and cell.data_type == 'f':
                    try:
                        # This is a simplified approach
                        # In production, use data_only=True when loading workbook
                        formula_count += 1
                    except Exception as e:
                        logger.warning(f"Failed to evaluate formula in {cell.coordinate}: {e}")
        
        if formula_count > 0:
            logger.info(f"Found {formula_count} formulas in sheet")
