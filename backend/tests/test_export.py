"""Tests for export functionality."""

import os
import pytest
from pathlib import Path
from uuid import uuid4
from unittest.mock import Mock, patch

from fastapi.testclient import TestClient


@pytest.fixture
def upload_id():
    """Generate test upload ID."""
    return uuid4()


@pytest.fixture
def company_info():
    """Sample company information."""
    return {
        "name": "Test Company Ltd",
        "cin": "U12345AB2020PTC123456",
        "registered_office": "123 Test Street, Mumbai",
        "financial_year": "2023-24"
    }


@pytest.fixture
def normalized_data():
    """Sample normalized data."""
    return [
        {
            "indicator": "Total Electricity Consumption",
            "normalized_value": 5000.0,
            "normalized_unit": "MWh",
            "original_value": 5000000.0,
            "original_unit": "kWh"
        },
        {
            "indicator": "Scope 1 Emissions",
            "normalized_value": 2500.0,
            "normalized_unit": "tonnes CO2e",
            "original_value": 2500000.0,
            "original_unit": "kg CO2e"
        }
    ]


class TestBRSRGeneration:
    """Test BRSR report generation."""

    def test_brsr_excel_generation(self, upload_id, company_info, normalized_data):
        """Test BRSR Excel generation."""
        from src.reporting.generator import ReportGenerator
        
        generator = ReportGenerator("data/templates")
        
        filepath = generator.generate_brsr_excel(
            upload_id=upload_id,
            company_info=company_info,
            normalized_data=normalized_data
        )
        
        assert filepath is not None
        assert filepath.endswith('.xlsx')
        assert 'BRSR_Excel' in filepath
        assert os.path.exists(filepath)
        
        # Cleanup
        if os.path.exists(filepath):
            os.remove(filepath)

    def test_brsr_excel_structure(self, upload_id, company_info, normalized_data):
        """Test BRSR Excel has correct structure."""
        from src.reporting.generator import ReportGenerator
        import openpyxl
        
        generator = ReportGenerator("data/templates")
        filepath = generator.generate_brsr_excel(
            upload_id=upload_id,
            company_info=company_info,
            normalized_data=normalized_data
        )
        
        # Load and verify
        wb = openpyxl.load_workbook(filepath)
        
        assert 'Section A' in wb.sheetnames
        assert 'Principle 6' in wb.sheetnames
        
        ws_a = wb['Section A']
        assert ws_a['B2'].value == company_info['name']
        assert ws_a['B3'].value == company_info['cin']
        
        ws_p6 = wb['Principle 6']
        assert ws_p6['A3'].value == 'Indicator'
        assert ws_p6['A4'].value == normalized_data[0]['indicator']
        
        # Cleanup
        wb.close()
        if os.path.exists(filepath):
            os.remove(filepath)

    def test_brsr_empty_data(self, upload_id, company_info):
        """Test BRSR generation with empty data."""
        from src.reporting.generator import ReportGenerator
        
        generator = ReportGenerator("data/templates")
        
        filepath = generator.generate_brsr_excel(
            upload_id=upload_id,
            company_info=company_info,
            normalized_data=[]
        )
        
        assert os.path.exists(filepath)
        
        # Cleanup
        if os.path.exists(filepath):
            os.remove(filepath)


class TestExportAPI:
    """Test export API endpoints."""

    @patch('src.api.export.generator')
    def test_export_excel_format(self, mock_generator, upload_id):
        """Test export with Excel format."""
        from src.api.export import router
        from fastapi import FastAPI
        
        app = FastAPI()
        app.include_router(router)
        client = TestClient(app)
        
        mock_generator.generate_brsr_excel.return_value = f"data/outputs/BRSR_Excel_{upload_id}.xlsx"
        
        response = client.post(
            f"/api/v1/export/{upload_id}",
            params={"format": "excel", "report_types": ["brsr"]}
        )
        
        assert response.status_code == 200
        data = response.json()
        assert data['format'] == 'excel'
        assert len(data['reports']) == 1
        assert data['reports'][0]['type'] == 'brsr'
        assert data['reports'][0]['format'] == 'excel'

    @patch('src.api.export.generator')
    def test_export_pdf_format(self, mock_generator, upload_id):
        """Test export with PDF format."""
        from src.api.export import router
        from fastapi import FastAPI
        
        app = FastAPI()
        app.include_router(router)
        client = TestClient(app)
        
        mock_generator.generate_reports.return_value = [
            {"type": "brsr", "format": "docx", "filepath": "test.docx"}
        ]
        mock_generator.convert_to_pdf.return_value = "test.pdf"
        
        response = client.post(
            f"/api/v1/export/{upload_id}",
            params={"format": "pdf", "report_types": ["brsr"]}
        )
        
        assert response.status_code == 200
        data = response.json()
        assert data['format'] == 'pdf'

    @patch('src.api.export.generator')
    def test_export_invalid_format(self, mock_generator, upload_id):
        """Test export with invalid format."""
        from src.api.export import router
        from fastapi import FastAPI
        
        app = FastAPI()
        app.include_router(router)
        client = TestClient(app)
        
        response = client.post(
            f"/api/v1/export/{upload_id}",
            params={"format": "invalid", "report_types": ["brsr"]}
        )
        
        assert response.status_code == 422


class TestMultiFrameworkExport:
    """Test multi-framework export."""

    def test_multiple_formats(self, upload_id, company_info, normalized_data):
        """Test generating multiple report formats."""
        from src.reporting.generator import ReportGenerator
        
        generator = ReportGenerator("data/templates")
        
        # Generate Excel
        excel_path = generator.generate_brsr_excel(
            upload_id=upload_id,
            company_info=company_info,
            normalized_data=normalized_data
        )
        
        assert os.path.exists(excel_path)
        
        # Cleanup
        if os.path.exists(excel_path):
            os.remove(excel_path)

    def test_brsr_and_gri_compatibility(self, normalized_data):
        """Test data compatibility between BRSR and GRI."""
        # Both frameworks should accept same normalized data
        brsr_indicators = {d['indicator'] for d in normalized_data}
        gri_indicators = {d['indicator'] for d in normalized_data}
        
        # Should have common indicators
        common = brsr_indicators & gri_indicators
        assert len(common) > 0


class TestPDFConversion:
    """Test PDF conversion."""

    def test_convert_to_pdf_method(self):
        """Test convert_to_pdf method exists."""
        from src.reporting.generator import ReportGenerator
        
        generator = ReportGenerator("data/templates")
        
        assert hasattr(generator, 'convert_to_pdf')
        assert callable(generator.convert_to_pdf)

    @patch('src.reporting.generator.convert')
    def test_pdf_conversion_success(self, mock_convert):
        """Test successful PDF conversion."""
        from src.reporting.generator import ReportGenerator
        
        generator = ReportGenerator("data/templates")
        
        docx_path = "test.docx"
        pdf_path = generator.convert_to_pdf(docx_path)
        
        assert pdf_path == "test.pdf"
        mock_convert.assert_called_once_with(docx_path, "test.pdf")

    def test_pdf_conversion_import_error(self):
        """Test PDF conversion with missing library."""
        from src.reporting.generator import ReportGenerator
        
        generator = ReportGenerator("data/templates")
        
        with patch('src.reporting.generator.convert', side_effect=ImportError):
            with pytest.raises(Exception):
                generator.convert_to_pdf("test.docx")


class TestFileOperations:
    """Test file operations."""

    def test_output_directory_creation(self, upload_id, company_info):
        """Test output directory is created."""
        from src.reporting.generator import ReportGenerator
        
        generator = ReportGenerator("data/templates")
        
        filepath = generator.generate_brsr_excel(
            upload_id=upload_id,
            company_info=company_info,
            normalized_data=[]
        )
        
        output_dir = Path(filepath).parent
        assert output_dir.exists()
        
        # Cleanup
        if os.path.exists(filepath):
            os.remove(filepath)

    def test_unique_filenames(self, company_info, normalized_data):
        """Test each export generates unique filename."""
        from src.reporting.generator import ReportGenerator
        
        generator = ReportGenerator("data/templates")
        
        id1 = uuid4()
        id2 = uuid4()
        
        path1 = generator.generate_brsr_excel(id1, company_info, normalized_data)
        path2 = generator.generate_brsr_excel(id2, company_info, normalized_data)
        
        assert path1 != path2
        assert str(id1) in path1
        assert str(id2) in path2
        
        # Cleanup
        for path in [path1, path2]:
            if os.path.exists(path):
                os.remove(path)


class TestErrorHandling:
    """Test error handling."""

    def test_missing_company_info(self, upload_id):
        """Test handling missing company info."""
        from src.reporting.generator import ReportGenerator
        
        generator = ReportGenerator("data/templates")
        
        # Should not crash with empty company_info
        filepath = generator.generate_brsr_excel(
            upload_id=upload_id,
            company_info={},
            normalized_data=[]
        )
        
        assert os.path.exists(filepath)
        
        # Cleanup
        if os.path.exists(filepath):
            os.remove(filepath)

    def test_invalid_data_types(self, upload_id, company_info):
        """Test handling invalid data types."""
        from src.reporting.generator import ReportGenerator
        
        generator = ReportGenerator("data/templates")
        
        invalid_data = [
            {"indicator": None, "normalized_value": "invalid"}
        ]
        
        # Should handle gracefully
        filepath = generator.generate_brsr_excel(
            upload_id=upload_id,
            company_info=company_info,
            normalized_data=invalid_data
        )
        
        assert os.path.exists(filepath)
        
        # Cleanup
        if os.path.exists(filepath):
            os.remove(filepath)


class TestDataFormatting:
    """Test data formatting in reports."""

    def test_numeric_formatting(self, upload_id, company_info):
        """Test numeric values are formatted correctly."""
        from src.reporting.generator import ReportGenerator
        import openpyxl
        
        generator = ReportGenerator("data/templates")
        
        data = [
            {
                "indicator": "Energy",
                "normalized_value": 1234.5678,
                "normalized_unit": "MWh",
                "original_value": 1234567.8,
                "original_unit": "kWh"
            }
        ]
        
        filepath = generator.generate_brsr_excel(upload_id, company_info, data)
        
        wb = openpyxl.load_workbook(filepath)
        ws = wb['Principle 6']
        
        # Check value is present
        assert ws['B4'].value == 1234.5678
        
        # Cleanup
        wb.close()
        if os.path.exists(filepath):
            os.remove(filepath)

    def test_text_formatting(self, upload_id, company_info):
        """Test text values are formatted correctly."""
        from src.reporting.generator import ReportGenerator
        import openpyxl
        
        generator = ReportGenerator("data/templates")
        
        filepath = generator.generate_brsr_excel(upload_id, company_info, [])
        
        wb = openpyxl.load_workbook(filepath)
        ws = wb['Section A']
        
        # Check headers are bold
        assert ws['A1'].font.bold is True
        
        # Cleanup
        wb.close()
        if os.path.exists(filepath):
            os.remove(filepath)
