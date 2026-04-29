"""Pytest configuration and shared fixtures"""
import tempfile
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from src.common.database import get_db
from src.common.models import Base
from src.main import app

from tests.auth_helpers import attach_mock_auth_user


@pytest.fixture(scope="function")
def test_db():
    """Create in-memory SQLite database for testing"""
    engine = create_engine("sqlite:///:memory:", echo=False)
    Base.metadata.create_all(engine)
    
    TestingSessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
    db = TestingSessionLocal()
    
    try:
        yield db
    finally:
        db.close()
        Base.metadata.drop_all(engine)


@pytest.fixture(scope="function")
def client(test_db):
    """Create FastAPI test client with test database"""
    def override_get_db():
        try:
            yield test_db
        finally:
            pass
    
    app.dependency_overrides[get_db] = override_get_db
    attach_mock_auth_user(app)

    with TestClient(app) as test_client:
        yield test_client

    app.dependency_overrides.clear()


@pytest.fixture
def sample_excel_file():
    """Create sample Excel file with formulas"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    
    # Headers
    ws['A1'] = "Metric"
    ws['B1'] = "Value"
    ws['C1'] = "Calculated"
    
    # Data with formulas
    ws['A2'] = "Energy"
    ws['B2'] = 100
    ws['C2'] = "=B2*2"
    
    ws['A3'] = "Water"
    ws['B3'] = 200
    ws['C3'] = "=B3*3"
    
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        wb.save(tmp.name)
        yield tmp.name
    
    Path(tmp.name).unlink(missing_ok=True)


@pytest.fixture
def sample_excel_merged():
    """Create Excel file with merged cells"""
    wb = Workbook()
    ws = wb.active
    
    # Create merged cells
    ws.merge_cells('A1:B1')
    ws['A1'] = "Merged Header"
    
    ws['A2'] = "Data1"
    ws['B2'] = "Data2"
    
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        wb.save(tmp.name)
        yield tmp.name
    
    Path(tmp.name).unlink(missing_ok=True)


@pytest.fixture
def sample_csv_comma():
    """Create CSV file with comma delimiter"""
    with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False) as tmp:
        tmp.write("Name,Age,City\n")
        tmp.write("Alice,30,NYC\n")
        tmp.write("Bob,25,LA\n")
        tmp.flush()
        yield tmp.name
    
    Path(tmp.name).unlink(missing_ok=True)


@pytest.fixture
def sample_csv_semicolon():
    """Create CSV file with semicolon delimiter"""
    with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False) as tmp:
        tmp.write("Name;Age;City\n")
        tmp.write("Alice;30;NYC\n")
        tmp.write("Bob;25;LA\n")
        tmp.flush()
        yield tmp.name
    
    Path(tmp.name).unlink(missing_ok=True)


@pytest.fixture
def sample_csv_tab():
    """Create CSV file with tab delimiter"""
    with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False) as tmp:
        tmp.write("Name\tAge\tCity\n")
        tmp.write("Alice\t30\tNYC\n")
        tmp.write("Bob\t25\tLA\n")
        tmp.flush()
        yield tmp.name
    
    Path(tmp.name).unlink(missing_ok=True)
